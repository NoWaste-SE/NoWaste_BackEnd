from django.test import TestCase, Client
from django.urls import reverse
# from django.contrib.auth import get_user_model
# from rest_framework.test import APITestCase, APIRequestFactory
# from rest_framework.authtoken.models import Token
# from rest_framework import status
# from User.views import OrderViewSet2
from User.models import *
# from rest_framework_simplejwt.tokens import AccessToken
from .models import *
from rest_framework.test import APIClient,APITestCase
from rest_framework import status
from datetime import datetime
import openpyxl
from django.core.files.base import ContentFile
from User.tests import UserActionsTestCase



# # creat test for orderviewset2
# class OrderViewSet2Test(APITestCase):
#     def setUp(self):
#         self.user = MyAuthor.objects.create(
#             email='test@gmail.com',
#             name='test',
#             is_staff=False,
#             is_active=True,
#             is_superuser=False,
#             is_admin=False,
#             password='1234',
#             role='customer'
#         )
#         access_token = AccessToken.for_user(self.user)
#         self.token = str(access_token)
#         self.factory = APIRequestFactory()
#         self.view = OrderViewSet2.as_view({'get': 'list'})
    
#     def test_order_list(self):
#         request = self.factory.get('/orders/')
#         request.headers = {'Authorization': 'Bearer ' + self.token}
#         response = self.view(request)
#         print(response)
#         self.assertEqual(response.status_code, status.HTTP_200_OK)

class FoodModelTest(TestCase):
    def setUp(self):
        manager = RestaurantManager.objects.create(
            email='manager@example.com',
            name='Test Manager',
            password='managerpassword',
            role='restaurant'
        )
        restaurant = Restaurant.objects.create(
            type='Iranian',
            address='Test Address',
            name='Test Restaurant',
            manager=manager
        )
        self.food = Food.objects.create(
            name='Test Food',
            price=10.99,
            ingredients='Ingredient 1, Ingredient 2',
            restaurant=restaurant,
            remainder=50
        )

    def test_food_model_str(self):
        self.assertEqual(str(self.food), 'Test Food')

    def test_save_method_default_food_pic(self):
        food = Food.objects.create(
            name='Food with Default Pic',
            price=15.99,
            ingredients='Ingredient 3, Ingredient 4',
            restaurant=self.food.restaurant,
            remainder=20
        )
        self.assertEqual(food.food_pic, default_food_pic)

    def test_save_method_default_food_pic2(self):
        food = Food.objects.create(
            name='Food with None Pic2',
            price=19.99,
            ingredients='Ingredient 5, Ingredient 6',
            restaurant=self.food.restaurant,
            remainder=30
        )
        self.assertIsNone(food.food_pic2)

    def test_save_method_no_default_food_pic(self):
        food = Food.objects.create(
            name='Food without Default Pic',
            price=25.99,
            ingredients='Ingredient 7, Ingredient 8',
            restaurant=self.food.restaurant,
            remainder=40,
            food_pic='custom_food_pic_value'
        )
        self.assertEqual(food.food_pic, 'custom_food_pic_value')

    def test_price_field(self):
        self.assertEqual(self.food.price, 10.99)

    def test_remainder_field(self):
        self.assertEqual(self.food.remainder, 50)

    def test_no_default_values(self):
        food = Food.objects.create(
            name='Custom Food',
            price=18.75,
            ingredients='Ingredient 13, Ingredient 14',
            restaurant=self.food.restaurant,
            remainder=25,
            food_pic='custom_food_pic',
            food_pic2='custom_food_pic2'
        )
        self.assertEqual(food.food_pic, 'custom_food_pic')
        self.assertEqual(food.food_pic2, 'custom_food_pic2')

    def test_empty_reminder(self):
        food = Food.objects.create(
            name='',
            ingredients='',
            restaurant=self.food.restaurant
        )
        self.assertEqual(food.remainder, 0) 


    def test_food_creation_with_long_name(self):
        long_name = 'A' * 256
        food_long_name = Food(
            name=long_name,
            price=10.00,
            ingredients='Ingredient 23, Ingredient 24',
            restaurant=self.food.restaurant,
            remainder=50
        )
        with self.assertRaises(Exception):
            food_long_name.save()

    def test_food_creation_with_long_ingredients(self):
        long_ingredients = 'A' * 2049
        food_long_ingredients = Food(
            name='Long Ingredients Food',
            price=12.00,
            ingredients=long_ingredients,
            restaurant=self.food.restaurant,
            remainder=60
        )
        with self.assertRaises(Exception):
            food_long_ingredients.save()

    def test_food_creation_with_large_price(self):
        large_price = 10**20
        food_large_price = Food(
            name='Large Price Food',
            price=large_price,
            ingredients='Ingredient 25, Ingredient 26',
            restaurant=self.food.restaurant,
            remainder=70
        )
        with self.assertRaises(Exception):
            food_large_price.save()

    def test_food_creation_with_large_remainder(self):
        large_remainder = 10**20
        food_large_remainder = Food(
            name='Large Remainder Food',
            price=15.00,
            ingredients='Ingredient 27, Ingredient 28',
            restaurant=self.food.restaurant,
            remainder=large_remainder
        )
        with self.assertRaises(Exception):
            food_large_remainder.save()


class OrderHistoryManagerExportExcelTestCase(TestCase):
    def setUp(self):
        self.manager = RestaurantManager.objects.create(name='Test Manager', email='test_manager@example.com')
        self.restaurant = Restaurant.objects.create(name='Test Restaurant', number='123', manager=self.manager)
        self.food = Food.objects.create(name='Test Food', price=10.0, restaurant=self.restaurant)
        self.user = Customer.objects.create(name='Test User', email='test_user@example.com')
        self.order = Order.objects.create(userId=self.user, status='Delivered', restaurant=self.restaurant)
        OrderItem.objects.create(order=self.order, food=self.food, quantity=2)

    def test_export_excel(self):
        client = APIClient()
        url = reverse('csv-order-history-manager', kwargs={'manager_id': self.manager.id})
        response = client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue('Content-Disposition' in response)
        # Check the filename
        expected_filename = 'restaurants-orders.xlsx'
        content_disposition = response['Content-Disposition']
        self.assertTrue(expected_filename in content_disposition)
        content_file = ContentFile(response.content)
        wb = openpyxl.load_workbook(content_file)
        ws = wb.active
        # Check the headers
        expected_headers = ['Restaurant Name', 'Restaurant Number', 'Food Name', 'Food Price', 'Quantity', 'User Name', 'User Email', 'Date', 'Status']
        for col_num, header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_num).value
            self.assertEqual(cell_value, header)
        # Check the data
        cell_values = [ws.cell(row=2, column=col_num).value for col_num in range(1, 10)]
        expected_data = [
            str(self.restaurant.name),
            str(self.restaurant.number),
            str(self.food.name),
            str("10.00"),
            str(2),
            str(self.user.name),
            str(self.user.email),
            str(self.order.created_at.date()), 
            str(self.order.status),
        ]
        self.assertEqual(cell_values, expected_data)



class OrderHistoryCustomerExportExcelTestCase(TestCase):
    def setUp(self):
        self.manager = RestaurantManager.objects.create(name='Test Manager', email='test_manager@example.com')
        self.restaurant = Restaurant.objects.create(name='Test Restaurant', number='123', manager=self.manager)
        self.food = Food.objects.create(name='Test Food', price=10, restaurant=self.restaurant)
        self.user = Customer.objects.create(name='Test User', email='test_user@example.com')
        self.order = Order.objects.create(userId=self.user, status='Delivered', restaurant=self.restaurant)
        OrderItem.objects.create(order=self.order, food=self.food, quantity=2)
    def test_export_excel(self):
        client = APIClient()
        url = reverse('csv-order-history-customer', kwargs={'restaurant_id': self.restaurant.id, 'userId': self.user.id})
        response = client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue('Content-Disposition' in response)
        expected_filename = 'orders-history.xlsx'
        content_disposition = response['Content-Disposition']
        self.assertTrue(expected_filename in content_disposition)
        content_file = ContentFile(response.content)
        wb = openpyxl.load_workbook(content_file)
        ws = wb.active
        # Check the headers
        expected_headers = ['Restaurant Name', 'Restaurant Number', 'Food Name', 'Food Price', 'Quantity', 'Date', 'Status']
        for col_num, header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_num).value
            self.assertEqual(cell_value, header)
        # Check the data
        cell_values = [ws.cell(row=2, column=col_num).value for col_num in range(1, 8)]
        expected_data = [
            str(self.restaurant.name),
            str(self.restaurant.number),
            str(self.food.name),
            str("10.00"),
            str(2), 
            str(self.order.created_at.date()),
            str(self.order.status),
        ]
        self.assertEqual(cell_values, expected_data)


class OrderHistoryDiffRestaurantCustomerExportExcelTestCase(TestCase):
    def setUp(self):
        self.manager = RestaurantManager.objects.create(name='Test Manager', email='test_manager@example.com')
        self.restaurant1 = Restaurant.objects.create(name='Restaurant 1', number='111', manager=self.manager)
        self.restaurant2 = Restaurant.objects.create(name='Restaurant 2', number='222', manager=self.manager)
        self.food1 = Food.objects.create(name='Food 1', price=10, restaurant=self.restaurant1)
        self.food2 = Food.objects.create(name='Food 2', price=15, restaurant=self.restaurant2)
        self.user = Customer.objects.create(name='Test User', email='test_user@example.com')
        self.order1 = Order.objects.create(userId=self.user, status='Delivered', restaurant=self.restaurant1)
        OrderItem.objects.create(order=self.order1, food=self.food1, quantity=2)
        self.order2 = Order.objects.create(userId=self.user, status='Pending', restaurant=self.restaurant2)
        OrderItem.objects.create(order=self.order2, food=self.food2, quantity=1)

    def test_export_excel(self):
        client = APIClient()
        url = reverse('csv-diff-order-history-customer', kwargs={'userId': self.user.id})
        response = client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue('Content-Disposition' in response)
        expected_filename = 'orders-history.xlsx'
        content_disposition = response['Content-Disposition']
        self.assertTrue(expected_filename in content_disposition)
        content_file = ContentFile(response.content)
        wb = openpyxl.load_workbook(content_file)
        ws = wb.active
        # Check the headers
        expected_headers = ['Restaurant Name', 'Restaurant Number', 'Food Name', 'Food Price', 'Quantity', 'Date', 'Status']
        for col_num, header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_num).value
            self.assertEqual(cell_value, header)
        # Check the data
        cell_values = [ws.cell(row=row, column=col).value for row in range(2, 4) for col in range(1, 8)]
        expected_data = [
            str(self.restaurant1.name),
            str(self.restaurant1.number),
            str(self.food1.name),
            str("10.00"),
            str(2), 
            str(self.order1.created_at.date()), 
            str(self.order1.status),
            str(self.restaurant2.name),
            str(self.restaurant2.number),
            str(self.food2.name),
            str("15.00"),
            str(1), 
            str(self.order2.created_at.date()),
            str(self.order2.status),
        ]
        self.assertEqual(cell_values, expected_data)


# class CommentModelTestCase(TestCase):
#     def setUp(self):
#         self.customer = Customer.objects.create(name='Test User', email='test_user@example.com')
#         self.manager = RestaurantManager.objects.create(name='Test Manager', email='test_manager@example.com')
#         self.restaurant = Restaurant.objects.create(name='Restaurant 1', number='111', manager=self.manager)
#         self.comment = Comment.objects.create(
#             restaurant=self.restaurant,
#             writer=self.customer,
#             text='This is a test comment'
#         )

#     def test_comment_str_representation(self):
#         self.assertEqual(str(self.comment), f'{self.customer.name} {self.restaurant.name}')

#     def test_comment_ordering(self):
#         later_comment = Comment.objects.create(
#             restaurant=self.restaurant,
#             writer=self.customer,
#             text='This is a later test comment'
#         )
#         comments = Comment.objects.all()
#         # self.assertEqual(comments[0], later_comment)
#         # self.assertEqual(comments[1], self.comment)

#     def test_comment_created_at_auto_now_add(self):
#         self.assertIsNotNone(self.comment.created_at)

#     def test_comment_default_text(self):
#         comment_without_text = Comment.objects.create(
#             restaurant=self.restaurant,
#             writer=self.customer
#         )
#         self.assertEqual(comment_without_text.text, '')

#     def test_comment_blank_text(self):
#         comment_with_blank_text = Comment.objects.create(
#             restaurant=self.restaurant,
#             writer=self.customer,
#             text=''
#         )
#         self.assertEqual(comment_with_blank_text.text, '')

#     def test_comment_foreign_keys(self):
#         self.assertEqual(self.comment.restaurant, self.restaurant)
#         self.assertEqual(self.comment.writer, self.customer)

#     def test_comment_update_text(self):
#         new_text = 'Updated test comment text'
#         self.comment.text = new_text
#         self.comment.save()
#         self.assertEqual(Comment.objects.get(pk=self.comment.pk).text, new_text)

#     def test_comment_delete(self):
#         comment_count_before = Comment.objects.count()
#         self.comment.delete()
#         self.assertEqual(Comment.objects.count(), comment_count_before - 1)


#     def test_comment_text_max_length(self):
#         long_text = 'A' * 1024
#         comment_long_text = Comment(
#             restaurant=self.restaurant,
#             writer=self.customer,
#             text = long_text
#         )
#         with self.assertRaises(Exception):
#             comment_long_text.save()

#     def test_comment_created_at_auto_now_add_accuracy(self):
#         current_time = datetime.now()
#         time_difference = timedelta(seconds=3)
#         with self.settings(USE_TZ=False):
#             comment = Comment.objects.create(
#                 restaurant=self.restaurant,
#                 writer=self.customer,
#                 text='This is a test comment'
#             )
#             self.assertTrue(current_time - time_difference <= comment.created_at <= current_time + time_difference)

# class CommentAPITestCase(APITestCase):
#     def setUp(self):
#         self.manager = RestaurantManager.objects.create(name='Test Manager', email='test_manager@example.com')
#         self.restaurant = Restaurant.objects.create(name='Restaurant 1', number='111', manager=self.manager)
#         self.customer_id = 0
#         self.url = ""
#     def athenticate(self, email, passw, name, role):
#         response = self.client.post(
#             reverse("signup"),
#             {
#                 "name": name,
#                 "email": email,
#             }
#         )
#         self.assertEqual(response.status_code, status.HTTP_201_CREATED)
#         response = self.client.post(
#             reverse("verify-email"),
#             {
#                 "name": name,
#                 "password": passw,
#                 "role": role,
#                 "email": email,
#                 "code": VC_Codes.objects.get(email=email).vc_code,
#             }
#         )      
#         self.assertEqual(response.status_code, status.HTTP_201_CREATED)
#         response = self.client.post(
#             reverse("login"),
#             {
#                 "email": email,
#                 "password": passw,
#             },
#         )
#         self.customer_id = response.data['id']
#         self.assertEqual(response.status_code, status.HTTP_200_OK)
#         token = response.data['access_token']
#         self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")

#     def test_post_comment(self):
#         self.athenticate('test_user@example.com', "test_pass", 'Test User', "customer") 
#         data = {'text': 'This is a test comment'}
#         user_id = self.customer_id
#         restaurant_id = self.restaurant.id
#         self.url = reverse('comment', kwargs={'restaurant_id': restaurant_id})
#         expected_url = f'/restaurant/comment/restaurant_id/{restaurant_id}/'
#         self.assertEqual(self.url, expected_url)
#         response = self.client.post(self.url, data)
#         self.assertEqual(response.status_code, status.HTTP_200_OK)
#         comment = Comment.objects.get(writer=self.customer_id, restaurant=self.restaurant)
#         self.assertEqual(comment.text, 'This is a test comment')

#     def test_get_comment(self):
#         self.athenticate('test_user@example.com', "test_pass", 'Test User', "customer")
#         user_id = self.customer_id
#         user = Customer.objects.get(id = self.customer_id)
#         restaurant_id = self.restaurant.id
#         self.url = reverse('comment', kwargs={'restaurant_id': restaurant_id})
#         comment = Comment.objects.create(writer=user, restaurant=self.restaurant, text='Test comment')
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, status.HTTP_200_OK)
#         self.assertEqual(response.data['comment'], 'Test comment')

#     def test_get_comment_nonexistent_comment(self):
#         self.athenticate('test_user@example.com', "test_pass", 'Test User', "customer")
#         user_id = self.customer_id
#         restaurant_id = self.restaurant.id
#         self.url = reverse('comment', kwargs={'restaurant_id': restaurant_id})
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, status.HTTP_200_OK)
#         self.assertIsNone(response.data.get('comment'))

#     def test_get_comment_invalid_user_or_restaurant(self):
#         response = self.client.get('/restaurant/comment/restaurant_id/invalid_restaurant_id/')
#         self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)



class FilterFoodViewSetTestCase(APITestCase):
    def setUp(self):
        self.manager = RestaurantManager.objects.create(name='Test Manager', email='test_manager@example.com')
        self.restaurant1 = Restaurant.objects.create(name='Restaurant1', number='111', manager=self.manager)
        self.restaurant2 = Restaurant.objects.create(name='Restaurant2', number='222', manager=self.manager)
        self.food1 = Food.objects.create(name='Food1', price=10, restaurant=self.restaurant1)
        self.food2 = Food.objects.create(name='Food2', price=15, restaurant=self.restaurant2)

    def test_filter_food_list(self):
        url = reverse('filter-food-list') 
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), Food.objects.count())

    def test_search_food_by_name(self):
        url = reverse('filter-food-list')
        response = self.client.get(url, {'search': 'Food1'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['name'], 'Food1')

    def test_order_food_list(self):
        url = reverse('filter-food-list')
        response = self.client.get(url, {'ordering': 'price'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        prices = [item['price'] for item in response.data]
        self.assertEqual(prices, sorted(prices))

    def test_search_food_case_insensitive(self):
        url = reverse('filter-food-list')
        response = self.client.get(url, {'search': 'food1'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['name'], 'Food1')

    def test_search_food_multiple_results(self):
        url = reverse('filter-food-list')
        response = self.client.get(url, {'search': 'Food'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), Food.objects.count())

    def test_order_food_list_descending(self):
        url = reverse('filter-food-list')
        response = self.client.get(url, {'ordering': '-price'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        prices = [item['price'] for item in response.data]
        self.assertEqual(prices, sorted(prices, reverse=True))

    def test_filter_food_by_price_range(self):
        url = reverse('filter-food-list')
        response = self.client.get(url, {'price__gte': 10.0, 'price__lte': 15.0})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        prices = [item['price'] for item in response.data]
        self.assertTrue(all(10.0 <= float(price) <= 15.0 for price in prices))



class RestaurantSearchViewSetTestCase(APITestCase):
    def setUp(self):
        self.manager = RestaurantManager.objects.create(name='Test Manager', email='test_manager@example.com')
        self.restaurant1 = Restaurant.objects.create(name='Restaurant1', number='111', manager=self.manager, rate=4.5, discount=0.1, date_of_establishment='2022-01-01')
        self.restaurant2 = Restaurant.objects.create(name='Restaurant2', number='222', manager=self.manager, rate=4.0, discount=0.05, date_of_establishment='2021-01-01')

    def test_search_restaurant_by_name(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'search': 'Restaurant1'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['name'], 'Restaurant1')

    def test_order_restaurant_list_by_rate(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'ordering': 'rate'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rates = [item['rate'] for item in response.data]
        self.assertEqual(rates, sorted(rates))

    def test_filter_restaurant_by_discount(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'discount__gte': 0.05, 'discount__lte': 0.1})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        discounts = [item['discount'] for item in response.data]
        self.assertTrue(all(0.05 <= float(discount) <= 0.1 for discount in discounts))

    def test_order_restaurant_list_by_date_of_establishment_descending(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'ordering': '-date_of_establishment'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        dates = [item['date_of_establishment'] for item in response.data]
        self.assertEqual(dates, sorted(dates, reverse=True))

    def test_search_restaurant_case_insensitive(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'search': 'restaurant1'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['name'], 'Restaurant1')

    def test_search_restaurant_multiple_results(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'search': 'Restaurant'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), Restaurant.objects.count())

    def test_filter_restaurant_by_rate_range(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'rate__gte': 4.0, 'rate__lte': 4.5})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rates = [item['rate'] for item in response.data]
        self.assertTrue(all(4.0 <= rate <= 4.5 for rate in rates))

    def test_order_restaurant_list_by_multiple_fields(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'ordering': ['rate', 'discount']})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rates_and_discounts = [(item['rate'], item['discount']) for item in response.data]
        self.assertEqual(rates_and_discounts, sorted(rates_and_discounts))

    def test_search_and_order_restaurants(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'search': 'Restaurant', 'ordering': 'rate'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), Restaurant.objects.count())
        rates = [item['rate'] for item in response.data]
        self.assertEqual(rates, sorted(rates))

    def test_search_and_filter_restaurants(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'search': 'Restaurant', 'discount__gte': 0.05})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        discounts = [item['discount'] for item in response.data]
        self.assertTrue(all(float(discount) >= 0.05 for discount in discounts))

    def test_search_order_and_filter_restaurants(self):
        url = reverse('restaurant-search-list')
        response = self.client.get(url, {'search': 'Restaurant', 'ordering': '-rate', 'discount__lte': 0.1})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        discounts = [item['discount'] for item in response.data]
        rates = [item['rate'] for item in response.data]
        self.assertEqual(rates, sorted(rates, reverse=True))
        self.assertTrue(all(float(discount) <= 0.1 for discount in discounts))