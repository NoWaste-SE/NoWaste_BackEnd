from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.viewsets import ModelViewSet 
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status ,generics,mixins,viewsets
from rest_framework import permissions
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.views import TokenVerifyView,TokenObtainPairView,TokenRefreshView
from rest_framework_simplejwt.authentication import JWTAuthentication
from .permissions import  IsAdminOrReadOnly
from .serializer import *
from .models import *
from .filters import RestaurantFilter , FoodFilter
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.renderers import JSONRenderer
from django.db.models import Q
import requests
import openpyxl
import csv
import json
from django.http import JsonResponse
import urllib
from rest_framework.renderers import JSONRenderer
from django.core import serializers
from profanity_check import predict, predict_prob
# from sklearn.externals import joblib

def GetUserByToken(request):
    authentication = JWTAuthentication()
    try:
        user, _ = authentication.authenticate(request)
    except Exception:
        user = None
    return user

'''class for Change Password API'''
class ChangePasswordView(generics.UpdateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = ChangePasswordSerializer
    def put(self, request, *args, **kwargs):
        user = request.user
        if(user.id != self.kwargs['pk']):
            return Response({"message": "Unathorized!"},status=status.HTTP_401_UNAUTHORIZED)
        super().update(request, *args, **kwargs) 
        return Response({"message" :"Password changed successfully!"},status= status.HTTP_200_OK)

'''class for accessing to List of restaurants or a specific restaurant and also update it .'''
class RestaurantProfileViewSet(viewsets.ViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    serializer_class = RestaurantSerializer

    def get_serializer(self, *args, **kwargs):
        serializer_class = self.serializer_class(*args, **kwargs)
        kwargs['context'] = self.get_serializer_context()
        return serializer_class
    
    def get_serializer_context(self):
        return {'request': self.request}
    
    # retrive specific restaurant
    def get_object(self, id):
        try:
            return Restaurant.objects.get(id=id)
        except Restaurant.DoesNotExist:
            raise Response(status= status.HTTP_404_NOT_FOUND)
        
    # return the list of all restaurants
    def list(self, request):
        queryset = Restaurant.objects.all()
        serializer = self.serializer_class(queryset, many=True)
        return Response(serializer.data)
    
    # retrieve specific restaurant and pass the results to the serializer
    def retrieve(self, request, id=None):
        queryset = Restaurant.objects.get(id=id)
        serializer = self.serializer_class(queryset)
        return Response(serializer.data)
    
    #update a specific restaurant
    def patch(self,request,id):
        instance = self.get_object(id= id)
        for key , value in request.data.items():
            setattr(instance,key,value)
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    # if there was an Id in request so we return the specific restaurant with that Id , else we return the list of all restaurants.
    def get(self, request, id=None):
        if id:
            return self.retrieve(request, id)
        else:
            return self.list(request)

'''class for retrieving specific restaurant to the customer'''
class RestaurantCustomerView(mixins.ListModelMixin,mixins.RetrieveModelMixin,viewsets.GenericViewSet):
    queryset = Restaurant.objects.all()
    # here we use specific serializer which show selected fields about restaurant 
    serializer_class = RestaurantSerializer
    lookup_field = 'id'
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def retrieve(self, request, id=None):
        user = GetUserByToken(request)
        queryset = Restaurant.objects.get(id=id)
        serializer = self.serializer_class(queryset)
        try:
            customer = Customer.objects.get(id=user.id)
            if RecentlyViewedRestaurant.objects.filter(user=customer, restaurant=queryset).exists():
                r = RecentlyViewedRestaurant.objects.get(user=customer, restaurant=queryset)
                r.delete()
            recently_view = RecentlyViewedRestaurant.objects.create(user=customer, restaurant=queryset)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except:
            return Response("User is not customer!", status=status.HTTP_400_BAD_REQUEST)

'''class for Food APIs'''
class FoodViewSet(ModelViewSet):
    serializer_class = FoodSerializer
    def get_queryset(self):
        print(self.kwargs)
        return Food.objects.filter(restaurant_id=self.kwargs['restaurant__id'])

    def get_serializer_context(self):
        return {'restaurant_id': self.kwargs['restaurant__id']}
    
    # class for updating a sepecific food
    def patch(self, request, id):
        instance = self.get_object(id= id)
        for key , value in request.data.items():
            setattr(instance,key,value)
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    
'''class for Listing foods of a Restaurant or adding to Restaurant's foods by its Restaurant manager'''
class ManagerFoodListCreateAPIView(generics.ListCreateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = FoodSerializer
    def get_queryset(self):
        print(self.kwargs)
        return Food.objects.filter(restaurant_id=self.kwargs['restaurant_id'])

    def get_serializer_context(self):
        print(self.kwargs)
        return {'restaurant_id': self.kwargs['restaurant_id']}

    def create(self, request, *args, **kwargs):
        serializer = FoodSerializer(data= request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def post(self, request, *args, **kwargs):
        return self.create(request, *args, **kwargs)

'''class for Restrive,Update,Destroy food of a Restaurant by its manager'''
class ManagerFoodViewSet(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = FoodSerializer
    lookup_field = 'pk'
    def get_queryset(self):
        print(self.kwargs)
        return Food.objects.filter(restaurant_id=self.kwargs['restaurant_id'])

    def get_serializer_context(self):
        print(self.kwargs)
        return {'restaurant_id': self.kwargs['restaurant_id']}

    def patch(self, request, id):
        instance = self.get_object(id= id)
        for key , value in request.data.items():
            setattr(instance,key,value)
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

'''class for searching a Restaurant'''
class RestaurantSearchViewSet(ModelViewSet):
    queryset = Restaurant.objects.all()
    serializer_class = RestaurantSearchSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = RestaurantFilter
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name']
    ordering_fields = ['rate', 'discount', 'name', 'date_of_establishment']
    def get_serializer_context(self):
        return {'request': self.request}
    
'''class for filtering food '''
class FilterFoodViewSet(ModelViewSet):
    queryset = Food.objects.all()
    serializer_class = FoodFilterSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = FoodFilter
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name']
    ordering_fields = ['price', 'name']
    def get_serializer_context(self):
        return {'request': self.request}

'''class for creating a Restaurant manager or get the list of managers'''
class RestaurantManagerListCreateView(generics.ListCreateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = RestaurantManager.objects.all()
    serializer_class = RestaurantManagerSerializer


'''class for Retrieve,Update,Destroy a Restaurant manager'''
class RestaurantManagerDetailView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = RestaurantManager.objects.all()
    serializer_class = RestaurantManagerSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['manager'] = self.get_object()
        return context
'''class for RestaurantManager API''' 
class RestaurantManagerRestaurantListView(generics.ListCreateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = Restaurant.objects.all()
    serializer_class = RestaurantSerializer

    def get_queryset(self):
        try:
            mang = RestaurantManager.objects.get(id=self.kwargs['manager_id'])
        except RestaurantManager.DoesNotExist:
            return Response("There is not any restaurant manager with the given id" , status=status.HTTP_404_NOT_FOUND)
        return self.queryset.filter(manager = mang)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['manager'] = RestaurantManager.objects.get(pk = self.kwargs['manager_id'])
        return context
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            try:
                mang = RestaurantManager.objects.get(id=self.kwargs['manager_id'])
            except RestaurantManager.DoesNotExist:
                return Response("There is not any restaurant manager with the given id" , status=status.HTTP_404_NOT_FOUND)
            serializer.save(manager = mang)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def post(self, request, *args, **kwargs):
        return self.create(request, *args, **kwargs)

'''class for Resturant manager Profile page'''        
class RestaurantManagerRestaurantDetailView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = Restaurant.objects.all()
    serializer_class = RestaurantSerializer # for show the manager's restaurants'information

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['manager'] = RestaurantManager.objects.get(pk=self.kwargs['manager_id'])
        return context

'''APIView for Orders'''
class OrderAPIView(generics.RetrieveUpdateAPIView,generics.CreateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = GetOrderSerializer
    lookup_field = 'pk'
    def get_queryset(self):
        return Order.objects.filter(restaurant_id=self.kwargs['restaurant_id'] ,userId_id = self.kwargs['userId'],status = "notOrdered").prefetch_related('orderItems').select_related('userId').select_related('restaurant')
    
    def get_serializer_class(self, *args, **kwargs):
        return GetOrderSerializer

    def get_serializer_context(self):
        return {'restaurant_id': self.kwargs['restaurant_id'],'userId_id' : self.kwargs['userId']}
    
    def retrieve(self, request, *args, **kwargs):
        instance =  self.get_queryset().filter(restaurant_id=self.kwargs['restaurant_id'],userId_id = self.kwargs['userId']).exclude(Q(status='Completed') | Q(status='Ordered'))
        serializer = None
        if (instance.count() == 0) :
            print("hiii")
            instance = Order.objects.create(restaurant_id=self.kwargs['restaurant_id'],userId_id = self.kwargs['userId'])
            serializer = self.get_serializer(instance)
        else :
            serializer = self.get_serializer(instance,many = True)
        return Response(serializer.data)
    
    def get(self, request, *args, **kwargs):
        return self.retrieve(request, *args, **kwargs)
    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)

'''API function for adding orderItem to order'''
def add_to_Order(request, *args, **kwargs):
    order  =  Order.objects.filter(restaurant_id=kwargs['restaurant_id'],userId_id = kwargs['userId'],status = 'notOrdered').first()
    instance = order
    if(order is None):
        # create new order and then add new orderitem
        try :
            order = Order.objects.create(restaurant_id=kwargs['restaurant_id'],userId_id = kwargs['userId'])
            instance = OrderItem.objects.create(food_id = kwargs['food_id'], order_id = order.id)
        # handle the exception
        except Exception as error:
            print("An exception occurred:", error) 
    else :
        # search that has an instance of the orderItem been added to the order befor?  
        instance = OrderItem.objects.filter(food_id = kwargs['food_id'], order_id = order.id).first()
        # if not, and the order has not had the same orderItem, so we should create new object for the new orderItem which is in request.
        if (instance is None):
            try :
                instance = OrderItem.objects.create(food_id = kwargs['food_id'], order_id = order.id)
            # if an error occured during the process of creating new orderitem, handle the exception
            except Exception as error:
                print("An exception occurred:", error)   
    # check the stock , if selected food is available(remainder>0) add to the queantity of orderItem.
    food = Food.objects.get(id = kwargs['food_id'])    
    try:  
        if food.remainder>0:
            instance.quantity = instance.quantity+ 1
            instance.save()  
            food.remainder -= 1
            food.save()
    except Exception as error:
        # handle the exception if any error ocured during the process
        print("An exception occurred:", error) 
    serializer = OrderItemSerializer(instance)
    serialized_data = serializer.data
    serialized_data['new_remainder'] = food.remainder

    content = JSONRenderer().render(serialized_data)
    return HttpResponse(content, content_type='application/json')

'''API function for removing orderItem from order'''
def remove_from_Order(request, *args, **kwargs):
    food = Food.objects.get(id = kwargs['food_id']) # find the orderItem object.(find the selected food to remove it of the order)
    order  =  Order.objects.filter(restaurant_id=kwargs['restaurant_id'],userId_id = kwargs['userId'],status = 'notOrdered').first()  # find the order according to this user and restaurant.
    instance = order
    if(order is None):
        data = {
            'name': food.name,
            'price': str(food.price),
            'message': "There is not any order between these user and restaurant",
            'new_remainder': food.remainder
        }
        json_data = json.dumps(data)
        return HttpResponse(json_data, content_type='application/json', status=status.HTTP_404_NOT_FOUND)
    else :
        instance = OrderItem.objects.filter(food_id = kwargs['food_id'], order_id = order.id).first()
        if (instance is None):
            data = {
            'name': food.name,
            'price': str(food.price),
            'message': "Customer didn't order this food",
            'new_remainder': food.remainder
            }
            json_data = json.dumps(data)
            return HttpResponse(json_data, content_type='application/json', status=status.HTTP_404_NOT_FOUND)
    try:    # substract one from the orderItem(instance) quantity
        instance.quantity = instance.quantity- 1
        instance.save()
        if (instance.quantity == 0):    # if the  orderItem quantity be Zero by substraction it means that there is no orderItem of this food in the order , any more.  
            food.remainder += 1     # one should added to the remainder stock,and then delete the orderitem from the order list.
            food.save()
            instance.delete()
            data = {
            'name': food.name,
            'price': str(food.price),
            'message': "The order item deleted from order",
            'new_remainder': food.remainder
            }
            json_data = json.dumps(data)
            return HttpResponse(json_data, content_type='application/json', status=status.HTTP_200_OK)
    except Exception as error:
    # handle the exception
        print("An exception occurred", error) 
    serializer = OrderItemSerializer(instance)
    serialized_data = serializer.data
    if instance.quantity > 0:   # add to the number of availables in stock by adding  one to the food.remainder 
        food.remainder += 1
        food.save()
    #update the new remainder in food serializer.
    serialized_data['new_remainder'] = food.remainder
    content = JSONRenderer().render(serialized_data)
    return HttpResponse(content, content_type='application/json')


'''class for Listing Customers' orders'''
class CustomerOrderViewAPI(generics.ListAPIView):
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated]
    def get_serializer_class(self):
        return CustomerViewOrderSerializer

    def get_queryset(self):
        return Order.objects.filter(userId_id=self.kwargs['user_id']).select_related('restaurant')

'''class for Listing Orders of a restaurant'''       
class RestaurantOrderViewAPI(generics.ListAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = RestaurantOrderViewSerializer
    def get_queryset(self):
        # Listing Orders of a restaurant(the resturanst are unified by their manager id)
        queryset = Restaurant.objects.filter(manager_id = self.kwargs['manager_id']).prefetch_related('orders')
        ordersList = []
        for restaurant in queryset:
            orders = restaurant.orders.all()
            for order in orders:
                ordersList.append(order)
        return ordersList

'''class for Updating the Order Status'''
class UpdateOrderStatusAPI(generics.UpdateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    lookup_url_kwarg = 'order_id'
    def get_serializer_class(self, *args, **kwargs):
        return UpdateOrderSerializer
    
    def get_queryset(self):
        return Order.objects.filter(id = self.kwargs['order_id'])
    
    def patch(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    
    def put(self, request, *args, **kwargs):
        return self.patch(request, *args, **kwargs)

'''class for post and get a comment'''
class CommentAPI(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        print("0000000000000000")
        serializer = CommentSerializer(data=request.data)
        writer = Customer.objects.get(id = kwargs['user_id'])
        restaurant = Restaurant.objects.get(id = kwargs['restaurant_id'])
        if serializer.is_valid(raise_exception=True):
            profanity_prediction = predict([serializer.validated_data['text']])[0]
            if profanity_prediction == 1:
                return Response({'error': 'Comment contains inappropriate content and cannot be saved.'}, status=status.HTTP_400_BAD_REQUEST)
            new_comment, created = Comment.objects.get_or_create(writer = writer, restaurant=restaurant)
            new_comment.text = serializer.validated_data['text']
            new_comment.save()
            return Response(serializer.validated_data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def get(self, request, *args, **kwargs):
        serializer = CommentSerializer()
        writer = Customer.objects.get(id = kwargs['user_id'])
        restaurant = Restaurant.objects.get(id = kwargs['restaurant_id'])
        try:
            comment = Comment.objects.get(writer=writer, restaurant=restaurant)
            return Response({'comment': comment.text}, status=status.HTTP_200_OK)
        except:
            return Response(serializer.data, status=status.HTTP_200_OK)

'''class for returning a list of comments according to a specific restaurant'''
class RestaurantCommentListAPIView(generics.ListAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = CommentSerializer
    queryset = Comment.objects.all()

    def get(self, request, *args, **kwargs):
        restaurant_id = self.kwargs['restaurant_id']
        comments = Comment.objects.filter(restaurant_id=restaurant_id)
        serializer = self.get_serializer(comments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

'''class for Search Nearest Restaurant from the map'''
class SearchNearestRestaurant(mixins.ListModelMixin):

    def search_nearest_restaurant(request):
        type_vehicle = 'car'
        origins = request.GET.get('origins')
        restaurants = Restaurant.objects.all()
        destinations = '|'.join([f"{restaurant.lat},{restaurant.lon}" for restaurant in restaurants])
        headers = {
            'Api-Key': 'service.f3f70682948d40999d64243013ff5b95',
        }
        url = f'https://api.neshan.org/v1/distance-matrix/no-traffic?type={type_vehicle}&origins={origins}&destinations={destinations}'
        response = requests.get(url,headers= headers)
        data = response.json()
        elements = data['rows'][0]['elements']
        destination_addresses = data['destination_addresses']
        dists = [element['distance']['value'] for element in elements]
        des_len = len(destination_addresses)
        des_dist_list = []
        for i in range(des_len):
            des_dist_list.append((elements[i]['distance']['value'],destination_addresses[i]))
        sorted_list = sorted(des_dist_list, key=lambda x: x[1])[:5]
        result = []
        for e in sorted_list:
            lat ,long = e[1].split(',')
            for rest in restaurants:
                if (rest.lat == lat,rest.lon == long):
                    result.append(rest)
        data = serializers.serialize('json', result)
        return HttpResponse(data, content_type="application/json")

'''a function API for retrieving address according to the lat and long'''
def get_addr(request):
    Lattitude = request.GET.get('lat')
    longitude = request.GET.get('lng')
    headers = {
        'Api-Key': 'service.7f461dfe908a40899d8900c2802f48a0',
    }
    url = f'https://api.neshan.org/v5/reverse?lat={Lattitude}&lng={longitude}'
    response = requests.get(url,headers= headers)
    data = response.json()
    return JsonResponse(data,safe= False)

'''updating the Latitude and logitude of the user's location'''
class LatLongUpdateRetreive(generics.RetrieveUpdateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = LatLongSerializer
    lookup_field = 'pk'
    lookup_url_kwarg = 'restaurant_id'
    def get_queryset(self):
        print(self.kwargs)
        return Restaurant.objects.filter(id=self.kwargs['restaurant_id'])

    def get_serializer_context(self):
        print(self.kwargs)
        return {'restaurant_id': self.kwargs['restaurant_id']}

    def patch(self, request, restaurant_id):
        instance = self.get_object(id= restaurant_id)
        for key , value in request.data.items():
            setattr(instance,key,value)
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

'''API class for returning the restaurant's lat and long'''
def get_lat_long(request, *args, **kwargs):
    rest = get_object_or_404(Restaurant,id =kwargs['restaurant_id'])
    content = JSONRenderer().render({'lat':rest.lat,'long':rest.lon})
    return HttpResponse(content, content_type='application/json')


def test(request):
    return HttpResponse("test")

'''download order history of all manager's restaurants as excel'''
class OrderHistoryManagerExportExcel(APIView):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="restaurants-orders.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['Restaurant Name', 'Restaurant Number', 'Food Name', 'Food Price', 'Quantity', 'User Name', 'User Email', 'Date', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Pink fill color
        queryset = Restaurant.objects.filter(manager_id=self.kwargs['manager_id']).prefetch_related('Orders__orderItems')
        row_num = 2
        for q in queryset:
            for o in q.Orders.all():
                for oi in o.orderItems.all():
                    ws.cell(row=row_num, column=1, value=str(q.name))
                    ws.cell(row=row_num, column=2, value=str(q.number))
                    ws.cell(row=row_num, column=3, value=str(oi.food.name))
                    ws.cell(row=row_num, column=4, value=str(oi.food.price))
                    ws.cell(row=row_num, column=5, value=str(oi.quantity))
                    ws.cell(row=row_num, column=6, value=str(o.userId.name))
                    ws.cell(row=row_num, column=7, value=str(o.userId.email))
                    ws.cell(row=row_num, column=8, value=str(o.created_at).split()[0])
                    ws.cell(row=row_num, column=9, value=str(o.status))
                    row_num += 1
        wb.save(response)
        return response

'''download order history for customer as excel'''
class OrderHistoryCustomerExportExcel(APIView):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="orders-history.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['Restaurant Name', 'Restaurant Number', 'Food Name', 'Food Price', 'Quantity', 'Date', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  
        queryset = Order.objects.filter(restaurant_id=self.kwargs['restaurant_id'], userId_id=self.kwargs['userId']).prefetch_related('orderItems').select_related('userId').select_related('restaurant')
        row_num = 2
        for o in queryset:
            for oi in o.orderItems.all():
                ws.cell(row=row_num, column=1, value=str(o.restaurant.name))
                ws.cell(row=row_num, column=2, value=str(o.restaurant.number))
                ws.cell(row=row_num, column=3, value=str(oi.food.name))
                ws.cell(row=row_num, column=4, value=str(oi.food.price))
                ws.cell(row=row_num, column=5, value=str(oi.quantity))
                ws.cell(row=row_num, column=6, value=str(o.created_at).split()[0])
                ws.cell(row=row_num, column=7, value=str(o.status))
                row_num += 1
        wb.save(response)
        return response

'''download order history of diffrent restaurants for customer as excel'''
class OrderHistoryDiffRestaurantCustomerExportExcel(APIView):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="orders-history.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['Restaurant Name', 'Restaurant Number', 'Food Name', 'Food Price', 'Quantity', 'Date', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  
        queryset = Order.objects.filter(userId_id=self.kwargs['userId']).prefetch_related('orderItems').select_related('userId').select_related('restaurant')
        row_num = 2
        for o in queryset:
            for oi in o.orderItems.all():
                ws.cell(row=row_num, column=1, value=str(o.restaurant.name))
                ws.cell(row=row_num, column=2, value=str(o.restaurant.number))
                ws.cell(row=row_num, column=3, value=str(oi.food.name))
                ws.cell(row=row_num, column=4, value=str(oi.food.price))
                ws.cell(row=row_num, column=5, value=str(oi.quantity))
                ws.cell(row=row_num, column=6, value=str(o.created_at).split()[0])
                ws.cell(row=row_num, column=7, value=str(o.status))
                row_num += 1
        wb.save(response)
        return response


class RecentlyViewedRestaurantsCusotmerView(APIView):
    queryset = RecentlyViewedRestaurant.objects.all()
    serializer_class = RecentlyViewedRestaurantSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = GetUserByToken(request)
        try:
            customer = Customer.objects.get(id=user.id)
            recently_views = RecentlyViewedRestaurant.objects.filter(user=customer).order_by('-viewed_at')[:6]
            serializer = self.serializer_class(recently_views, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except:
            return Response("Error!", status=status.HTTP_400_BAD_REQUEST)
