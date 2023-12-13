# https://testdriven.io/blog/django-custom-user-model/

# from django.contrib.auth import get_user_model
# from django.test import TestCase
from django.test import TestCase
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APIClient,APITestCase
from User.models import MyAuthor,VC_Codes
from .serializers import ForgotPasswordSerializer
from User.views import OrderViewSet2
from User.models import MyAuthor
from .models import *
from .serializers import *
import random , string
from django.db import connection
import openpyxl
from django.core.files.base import ContentFile


class UserActionsTestCase(APITestCase):

    custId = -1
    def signup(self,email,name):
        response = self.client.post(
            reverse("signup"),
            {
                "name": name,
                "email": email,
            }
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
    
    def verify_email(self,name,passw,role,email):    
        # Verify Email
        response = self.client.post(
            reverse("verify-email"),
            {
                "name": name,
                "password": passw,
                "role": role,
                "email": email,
                "code": VC_Codes.objects.get(email=email).vc_code,
            }
        )      
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
    def createSuperUser(self,email,passw):
        MyAuthor.objects.create_user(email = email,password = passw,is_admin = True,is_staff=True,is_superuser=True,role="admin")
    
    def get_customer_id(self):
        return self.custId

        # Login
    def login(self,email,passw):
        response = self.client.post(
            reverse("login"),
            {
                "email": email,
                "password": passw,
            },
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)        
        # Get Token
        token = response.data['access_token']
        self.custId = response.data['id']
        # self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        return token
    def signup_verifyEmail_login(self,email,passw,name,role):
        self.signup(email,name)
        self.verify_email(name,passw,role,email)
        token = self.login(email,passw)
        return token
    
class ForgotPasswordViewSetTests(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_forgot_password_valid_email(self):
        user = MyAuthor.objects.create(email='newuser@gmail.com', name='Test User')
        url = reverse('forgot-password') 
        data = {'email': 'newuser@gmail.com'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        # self.assertIn('to_email' in response.data, True)

    def test_forgot_password_invalid_email(self):
        url = reverse('forgot-password')
        data = {'email': 'invalid_email'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_forgot_password_user_not_found(self):
        url = reverse('forgot-password')
        data = {'email': 'nonexistent@example.com'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_forgot_password_successful_email_sent(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        with self.settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend'):
            url = reverse('forgot-password')  
            data = {'email': 'test@example.com'}
            response = self.client.post(url, data, format='json')
            self.assertEqual(response.status_code, status.HTTP_201_CREATED)
            # self.assertIn('to_email' in response.data, True)

    def test_forgot_password_invalid_method(self):
        url = reverse('forgot-password') 
        response = self.client.put(url)
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)


    def test_forgot_password_code_generation(self):
        user = MyAuthor.objects.create(email='n.haghighy@gmail.com', name='Test User')
        with self.settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend'):
            url = reverse('forgot-password')  
            data = {'email': 'n.haghighy@gmail.com'}
            response = self.client.post(url, data, format='json')
            self.assertEqual(response.status_code, status.HTTP_201_CREATED)
            vc_code = VC_Codes.objects.get(email='n.haghighy@gmail.com').vc_code
            self.assertTrue(vc_code.isalnum() and len(vc_code) == 10)

    def test_forgot_password_code_update_on_resend(self):
        user = MyAuthor.objects.create(email='n.haghighy@gmail.com', name='Test User')
        with self.settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend'):
            url = reverse('forgot-password') 
            data = {'email': 'n.haghighy@gmail.com'}
            response = self.client.post(url, data, format='json')
            self.assertEqual(response.status_code, status.HTTP_201_CREATED)
            initial_vc_code = VC_Codes.objects.get(email='n.haghighy@gmail.com').vc_code
            response_resend = self.client.post(url, data, format='json')
            self.assertEqual(response_resend.status_code, status.HTTP_201_CREATED)
            updated_vc_code = VC_Codes.objects.get(email='n.haghighy@gmail.com').vc_code
            self.assertNotEqual(initial_vc_code, updated_vc_code)


class ForgotPassVerifyTests(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_forgot_pass_verify_successful(self):
        user = MyAuthor.objects.create(email='test2@gmail.com', name='Test User')
        vc_code = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(10))
        VC_Codes.objects.create(email=user.email, name=user.name, vc_code=vc_code)
        url = reverse('fp-verify')  
        data = {'email': 'test2@gmail.com', 'code': vc_code}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_forgot_pass_verify_invalid_email(self):
        url = reverse('fp-verify')
        data = {'email': 'invalid_email', 'code': '1234567890'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_forgot_pass_verify_user_not_found(self):
        url = reverse('fp-verify')
        data = {'email': 'nonexistent@example.com', 'code': '1234567890'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_forgot_pass_verify_wrong_code(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        VC_Codes.objects.create(email=user.email, name=user.name, vc_code='1234567890')
        url = reverse('fp-verify')  
        data = {'email': 'test@example.com', 'code': '0987654321'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('verification code is wrong', response.data)

    def test_forgot_pass_verify_invalid_data(self):
        url = reverse('fp-verify')  
        data = {'email': 'test@example.com'}  # Missing 'code' field
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_forgot_pass_verify_code_deleted_after_successful_verification(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        vc_code = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(10))
        VC_Codes.objects.create(email=user.email, name=user.name, vc_code=vc_code)
        url = reverse('fp-verify') 
        data = {'email': 'test@example.com', 'code': vc_code}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        with self.assertRaises(VC_Codes.DoesNotExist):
            VC_Codes.objects.get(email='test@example.com')

    def test_forgot_pass_verify_multiple_attempts(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        vc_code = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(10))
        VC_Codes.objects.create(email=user.email, name=user.name, vc_code=vc_code)
        url = reverse('fp-verify')
        for _ in range(3):
            data = {'email': 'test@example.com', 'code': 'wrong_code'}
            response = self.client.post(url, data, format='json')
            self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
            self.assertIn('verification code is wrong', response.data)
        data = {'email': 'test@example.com', 'code': vc_code}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class ForgotPassSetNewPassTests(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_forgot_pass_set_new_pass_successful(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        url = reverse('fp-newpassword')  
        data = {'email': 'test@example.com', 'password': 'newpassword'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        user.refresh_from_db()
        self.assertTrue(user.check_password('newpassword'))

    def test_forgot_pass_set_new_pass_invalid_email(self):
        url = reverse('fp-newpassword')  
        data = {'email': 'invalid_email', 'password': 'newpassword'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_forgot_pass_set_new_pass_user_not_found(self):
        url = reverse('fp-newpassword')  
        data = {'email': 'nonexistent@example.com', 'password': 'newpassword'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_forgot_pass_set_new_pass_invalid_data(self):
        url = reverse('fp-newpassword')  
        data = {'email': 'test@example.com'}  # Missing 'password' field
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_forgot_pass_set_new_pass_password_update(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        old_password = user.password
        url = reverse('fp-newpassword') 
        data = {'email': 'test@example.com', 'password': 'newpassword'}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        user.refresh_from_db()
        self.assertNotEqual(user.password, old_password)

    def test_forgot_pass_set_new_pass_empty_password(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        url = reverse('fp-newpassword') 
        data = {'email': 'test@example.com', 'password': ''}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_forgot_pass_set_new_pass_whitespace_password(self):
        user = MyAuthor.objects.create(email='test@example.com', name='Test User')
        url = reverse('fp-newpassword') 
        data = {'email': 'test@example.com', 'password': '    '}
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class DatabaseConnectionTests(TestCase):
    def test_database_connection_is_open(self):
        # Check if the database connection is open
        self.assertTrue(connection.connection is not None)

    def test_database_connection(self):
        with connection.cursor() as cursor:
            cursor.execute('SELECT 1')
            result = cursor.fetchone()
        self.assertEqual(result, (1,))

    def test_database_name(self):
        with connection.cursor() as cursor:
            cursor.execute('SELECT current_database()')
            result = cursor.fetchone()
        self.assertEqual(result[0], 'test_nowaste')

    def test_create_and_retrieve_model(self):
        MyAuthor.objects.create(name='Test Item', email='test@example.com')
        # Retrieve the instance from the database
        test_item = MyAuthor.objects.get(name='Test Item')
        self.assertEqual(test_item.name, 'Test Item')
        self.assertEqual(test_item.email, 'test@example.com')

    def test_update_model(self):
        test_item = MyAuthor.objects.create(name='Test Item', email='test@example.com')
        test_item.name = 'Updated name'
        test_item.save()
        updated_item = MyAuthor.objects.get(email='test@example.com')
        self.assertEqual(updated_item.name, 'Updated name')

    def test_delete_model(self):
        MyAuthor.objects.create(name='Test Item', email='test@example.com')
        MyAuthor.objects.filter(name='Test Item').delete()
        deleted_item = MyAuthor.objects.filter(name='Test Item').first()
        self.assertIsNone(deleted_item)
        with self.assertRaises(MyAuthor.DoesNotExist):
            MyAuthor.objects.get(name='Test Item')

class OrderViewSet2TestCase(APITestCase):
    def setUp(self):
        self.url = reverse('order')
    
    def athenticate(self, email, passw, name, role):
        # SignUp
        response = self.client.post(
            reverse("signup"),
            {
                "name": name,
                "email": email,
            }
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        
        # Verify Email
        response = self.client.post(
            reverse("verify-email"),
            {
                "name": name,
                "password": passw,
                "role": role,
                "email": email,
                "code": VC_Codes.objects.get(email=email).vc_code,
            }
        )      
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        
        # Login
        response = self.client.post(
            reverse("login"),
            {
                "email": email,
                "password": passw,
            },
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Get Token
        token = response.data['access_token']
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
    def test_order_list(self):
        self.athenticate("test_email@gmail.com", "test_pass", "test_name", "customer")       
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
    def test_order_unathenticated_user(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
    
    def test_order_method_notAllowed(self):
        self.athenticate("test_email@gmail.com", "test_pass", "test_name", "customer")
        response = self.client.post(self.url, {})
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)


'''Implement LoginViewTest'''
class LoginViewTestCase(UserActionsTestCase):
    def setUp(self):
        self.url = reverse('login')
    
    def test_login_checkAuthentication(self):
        token = self.signup_verifyEmail_login("test_email@gmail.com", "test_pass", "test_name", "customer")       
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
    
    def test_login_method_notAllowed(self):
        self.signup_verifyEmail_login("test_email@gmail.com", "test_pass", "test_name", "customer")
        response = self.client.delete(self.url, {
                                                    "password": "test_pass",
                                                        "email": "test_email@gmail.com"
                                                    })
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)
        


class RestaurantInfoExportExcelTestCase(TestCase):
    def setUp(self):
        self.restaurant = Restaurant.objects.create(
            name='Test Restaurant',
            type='Iranian',
            address='Test Address',
            discount=0.3,
            rate=4.5,
            number='123456789',
            manager=RestaurantManager.objects.create(name='Test Manager', email='test@example.com')
        )

    def test_export_excel(self):
        client = APIClient()
        url = reverse('csv-restaurants-info')
        response = client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue('Content-Disposition' in response)
        # Check the file name
        expected_filename = 'restaurants-info.xlsx'
        content_disposition = response['Content-Disposition']
        self.assertTrue(expected_filename in content_disposition)
        content_file = ContentFile(response.content)
        wb = openpyxl.load_workbook(content_file)
        ws = wb.active
        # Check the headers
        expected_headers = ['Name', 'Type', 'Address', 'Discount', 'Rate', 'Number', 'Manager Name', 'Manager Email']
        for col_num, header in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=1, column=col_num).value
            self.assertEqual(cell_value, header)
        # Check the data
        cell_values = [ws.cell(row=2, column=col_num).value for col_num in range(1, 9)]
        expected_data = [self.restaurant.name, self.restaurant.type, self.restaurant.address,
                         self.restaurant.discount, self.restaurant.rate, self.restaurant.number,
                         self.restaurant.manager.name, self.restaurant.manager.email]
        self.assertEqual(cell_values, expected_data)

class AdminViewTestCase(UserActionsTestCase):
    def setUp(self):
        self.url = reverse('admin-profile')
    
    def test_adminPanel_checkAuthorization(self):
        token = self.signup_verifyEmail_login("test_email@gmail.com", "test_pass", "test_name", "customer")       
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_adminPanel_method_notAllowed(self):
        # self.athenticate("test_email@gmail.com", "test_pass", "test_name", "customer")
        self.createSuperUser("admin3@gmail.com","1234")
        token = self.login("admin3@gmail.com", "1234")
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        response = self.client.delete(self.url, {
                                                    "password": "test_pass",
                                                        "email": "test_email@gmail.com"
                                                    })
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)

class TempManagerAPITestCase(UserActionsTestCase):
    def setUp(self):
        # self.url = reverse('reject-tmpMng',kwargs={'pk':self.custId})
        self.url = reverse('confirm-tmpMng')
        self.tempManagerId = -1
    def creatTempManager(self,email,name):
        tmpMNG = TempManager.objects.create(email=email,name=name)
        self.tempManagerId = tmpMNG.id
    def test_managerConfirmation_checkAuthentication(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_managerConfirmation_checkAuthorization(self):
        token = self.signup_verifyEmail_login("test_email@gmail.com", "test_pass", "test_name", "customer")       
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class TempManagerConfirmationTestCase(TempManagerAPITestCase):
    def setUp(self):
        self.url = reverse('confirm-tmpMng')

    def test_managerConfirmation_method_notAllowed(self):
        # self.athenticate("test_email@gmail.com", "test_pass", "test_name", "customer")
        self.createSuperUser("admin3@gmail.com","1234")
        token = self.login("admin3@gmail.com", "1234")
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        tmpMNGemail = "temp@gmail.com"
        tmpMNGname= "tmpMNG1"
        self.creatTempManager(tmpMNGemail,tmpMNGname)
        response = self.client.delete(self.url, {
                                                        "email": tmpMNGemail,
                                                        "name":tmpMNGname
                                                     })
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)


class TempManagerRejectionTestCase(TempManagerAPITestCase):
    def setUp(self):
        self.creatTempManager("temp2@gmail.com","tmpMNG2")
        self.url = reverse('reject-tmpMng',kwargs={'pk':self.tempManagerId})

    def test_managerRejection_method_notAllowed(self):
        # self.athenticate("test_email@gmail.com", "test_pass", "test_name", "customer")
        self.createSuperUser("admin3@gmail.com","1234")
        token = self.login("admin3@gmail.com", "1234")
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        response = self.client.post(self.url,
                                                    {
                                                    "name": "tempMNG2",
                                                        "email": "temp3@gmail.com"
                                                    })
        # self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)
        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)

