from django.contrib.auth import logout
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework import status ,generics , permissions , mixins,viewsets
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .serializers import *
from .models import *
from .utils import Util
from rest_framework.authtoken.models import Token
from rest_framework.renderers import JSONRenderer
from django.template.loader import render_to_string
from django.core.validators import EmailValidator
from django.forms import ValidationError
import openpyxl
import random , string
import csv
import json
from cities_light.models import Country, City
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.views import TokenVerifyView,TokenObtainPairView,TokenRefreshView
from rest_framework_simplejwt.authentication import JWTAuthentication

'''Email Verification class for signing up'''
class VerifyEmail(APIView):
    def get_serializer_class(self, request):
        if request.data['role'] == "customer":
            return CreateCustomerSerializer
        elif request.data['role'] == "restaurant":
            return CreateRestaurantSerializer
    
    def post(self, request):
        serializer_class = self.get_serializer_class(request)
        serializer = serializer_class(data=request.data)
        if serializer.is_valid():
            user_data = serializer.validated_data
            try:
                user = VC_Codes.objects.get(email=user_data['email'])
            except VC_Codes.DoesNotExist:
                return Response("There is not any user with the given email" , status=status.HTTP_404_NOT_FOUND)
            if user_data['code'] == user.vc_code:
                VC_Codes.objects.filter(vc_code = user.vc_code).delete()   
                if user_data['role'] == "customer":
                    serializer.save()
                    myauthor = MyAuthor.objects.get(email = user_data['email'])
                    myauthor.role = user_data['role']
                    myauthor.save()
                    return Response(user_data, status=status.HTTP_201_CREATED)
                if user_data['role']=="restaurant":
                    if TempManager.objects.filter(email=user_data['email']).exists():
                        return Response("You are already waiting for admin approval", status=status.HTTP_400_BAD_REQUEST)
                    tmp = TempManager.objects.create(email=user_data['email'], name=user_data['name'])
                    tmp.set_password(user_data['password'])
                    tmp.save()
                    return Response("Please wait for admin confirmation.", status=status.HTTP_200_OK)

        return Response("verification code is wrong", status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        serializer = BaseCreateUserSerializer()
        return Response(serializer.data)
    
'''SignUp class'''   
class SignUpView(APIView):
    serializer_class = SignUpSerializer
    permission_classes = (permissions.AllowAny,)
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception= True) and not MyAuthor.objects.filter(email = request.data['email']).exists():
            vc_code = random.randint(100000, 999999)
            instance = serializer.save()
            instance.vc_code = vc_code
            instance.save()
            signupData = serializer.data
        else:
            return Response("email is already exist", status=status.HTTP_400_BAD_REQUEST)
        email = signupData['email']
        name = signupData['name']
        template = render_to_string('email_template.html', {'name': name, 'code': vc_code})
        data = {'to_email': email, 'body': template, 'subject': 'Welcome to NoWaste!(Verify your email)'}
        Util.send_email(data)
        return Response(signupData, status=status.HTTP_201_CREATED)
    def get(self,request):
        serializer = SignUpSerializer()
        return Response(serializer.data)
    
'''Login API view'''   
class LoginView(TokenObtainPairView):
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        access_token = response.data['access']
        refresh_token = response.data['refresh']
        if response.status_code == 200:
            email = request.data.get('email')
            password = request.data.get('password')
            try :
                user = MyAuthor.objects.get(email = email)

            except Exception as error :
                return Response({'error': 'Invalid email or password'}, status=status.HTTP_401_UNAUTHORIZED)
            if user is not None and user.check_password(password):
                id = user.id
                if user.role == "customer":
                    c = Customer.objects.get (email = email)
                    name = c.name
                    WalletBalance = c.wallet_balance
                    listOfFavorite = list(c.list_of_favorites_res.values_list('name', flat=True))
                    result_fav = []
                    for r in listOfFavorite:
                        res = Restaurant.objects.get(name = r)
                        result_fav.append({'address': res.address, 'name': res.name, 'restaurant_image': res.restaurant_image, 'discount': res.discount, 'number': res.number, 'rate': res.rate, 'date_of_establishment': res.date_of_establishment, 'description': res.description, 'id': res.id})
                    # listOfFavorite = list(c.list_of_favorites_res)
                    return Response({'access_token': access_token,'refresh_token':refresh_token,'id' : user.id, 'wallet_balance':WalletBalance, 'role':user.role, 'list_of_favorites_res':result_fav, 'name':name})
                else:
                    r = RestaurantManager.objects.get(email = email)
                    return Response({'access_token': access_token,'refresh_token':refresh_token,'id' : user.id, 'role':user.role, 'name':r.name})
            else:
                return Response({'error': 'Invalid email or password'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return response
    def get(self,request):
        serializer = MyAuthorSerializer()
        return Response(serializer.data)

'''Logout API view'''
class LogoutView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user
        Token.objects.filter(user=user).delete()
        logout(request)
        return Response({'message': 'User logged out successfully'})

'''Forgot Password API view'''
class ForgotPasswordViewSet(APIView):
    def post(self, request):
        serializer = ForgotPasswordSerializer()
        validate_email = EmailValidator()
        email = request.data.get('email')
        try:
            validate_email.__call__(email)
        except ValidationError as e:
            return Response(e.message, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = MyAuthor.objects.get(email=email)
        except MyAuthor.DoesNotExist:
            return Response("There is not any user with the given email" , status=status.HTTP_404_NOT_FOUND)
        newCode = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(10))
        try :
            u , created = VC_Codes.objects.get_or_create(email = user.email , name = user.name)
            u.vc_code = newCode
            u.save()
        except Exception as error:
            return HttpResponse(json.dumps({'error': error}), mimetype="application/json")
        template = render_to_string('forgotpass_template.html',
            {'name': u.name,
                'code': newCode})
        data = {'to_email':u.email,'body':template, 'subject': 'NoWaste forgot password'}
        Util.send_email(data)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    def get(self, request):
        serializer = ForgotPasswordSerializer()
        return Response(serializer.data)

'''Email Verification for Forgot Password '''
class ForgotPassVerify(APIView):
    def post(self, request):
        serializer = EmailVerificationSerializer(data=request.data)
        validate_email = EmailValidator()
        if serializer.is_valid():
            email =serializer.validated_data['email']
            try:
                validate_email.__call__(email)
            except ValidationError as e:
                return Response(e.message, status=status.HTTP_400_BAD_REQUEST)
            try:
                user = VC_Codes.objects.get(email=email)
            except VC_Codes.DoesNotExist:
                return Response("There is not any user with the given email" , status=status.HTTP_404_NOT_FOUND)
            if user.vc_code == serializer.validated_data['code']:
                new_serializer = MyAuthorSerializer()
                user.delete()
                return Response(new_serializer.data, status=status.HTTP_200_OK)
            else:
                return Response("verification code is wrong", status=status.HTTP_400_BAD_REQUEST)
        return Response("error!", status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        serializer = EmailVerificationSerializer()
        return Response(serializer.data)

'''Forgot Password new pass generating'''
class ForgotPassSetNewPass(APIView):
    def post(self, request):
        serializer = MyAuthorSerializer(data=request.data)
        validate_email = EmailValidator()
        if serializer.is_valid(raise_exception=True):
            email =serializer.validated_data['email']
            try:
                validate_email.__call__(email)
            except ValidationError as e:
                return Response(e.message, status=status.HTTP_400_BAD_REQUEST)
            try:
                user = MyAuthor.objects.get(email=email)
            except MyAuthor.DoesNotExist:
                return Response("There is not any user with the given email" , status=status.HTTP_404_NOT_FOUND)
            user.set_password(serializer.validated_data['password'])
            user.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response("error!", status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        serializer = MyAuthorSerializer()
        return Response(serializer.data)

'''Change Password API view '''
class ChangePasswordView(generics.UpdateAPIView):
    queryset = MyAuthor.objects.all()
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = ChangePasswordSerializer
    def put(self, request, *args, **kwargs):
        user = request.user
        if(user.id != self.kwargs['pk']):
            return Response({"message": "Unathorized!"},status=status.HTTP_401_UNAUTHORIZED)
        super().update(request, *args, **kwargs) 
        return Response({"message" :"Password changed successfully!"},status= status.HTTP_200_OK)

'''Profile update and Retrive API view '''
class UpdateRetrieveProfileView(generics.RetrieveUpdateAPIView):
    authentication_classes = [JWTAuthentication]  
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        return Customer.objects.filter(id=self.kwargs['id'])
    def get_serializer_class(self):
        if (self.request.method == 'GET'):
            return CustomerSerializer
        else :
            return UpdateUserSerializer
    lookup_field = 'id'
    def patch(self, request, *args, **kwargs):
        instance = self.get_object()
        for key , value in request.data.items():
            print(key , value)
            setattr(instance,key,value)
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    def get(self,request,id):
        if ( request.user.id != id ):
            return Response({"message": "Unathorized!"},status= status.HTTP_401_UNAUTHORIZED)
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

'''Rate Restaurant API'''
class RateRestaurantView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def post(self, request):
        serializer = RateRestaurantSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            name = serializer.validated_data['name']
            try:
                restaurant = Restaurant.objects.get(name=name)
            except Restaurant.DoesNotExist:
                return Response("There is not any restaurant with the given name" , status=status.HTTP_404_NOT_FOUND)
            restaurant.rate = round(((restaurant.rate) * restaurant.count_rates + serializer.validated_data['rate']) / (restaurant.count_rates + 1), 1)
            restaurant.count_rates += 1
            restaurant.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response("error!", status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        serializer = RateRestaurantSerializer()
        return Response(serializer.data)

'''Favorite Restaurant List API'''
class AddRemoveFavorite(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def post(self, request):
        serializer = AddRemoveFavoriteSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            name = serializer.validated_data['name']
            try:
                restaurant = Restaurant.objects.get(name=name)
            except Restaurant.DoesNotExist:
                return Response("There is not any restaurant with the given name" , status=status.HTTP_404_NOT_FOUND)
            email = serializer.validated_data['email']
            try:
                user = Customer.objects.get(email=email)
            except Customer.DoesNotExist:
                return Response("There is not any customer with the given email" , status=status.HTTP_404_NOT_FOUND)
            if user.list_of_favorites_res.filter(pk=restaurant.pk).exists():
                user.list_of_favorites_res.remove(restaurant)
            else:
                user.list_of_favorites_res.add(restaurant)
            user.save()
            listOfFavorite = list(user.list_of_favorites_res.values_list('name', flat=True))
            result_fav = []
            for r in listOfFavorite:
                res = Restaurant.objects.get(name = r)
                result_fav.append({'address': res.address, 'name': res.name, 'restaurant_image': res.restaurant_image, 'discount': res.discount, 'number': res.number, 'rate': res.rate, 'date_of_establishment': res.date_of_establishment, 'description': res.description, 'id': res.id})
            return Response({'list_of_favorites_res':result_fav}, status=status.HTTP_200_OK)
        return Response("Error!", status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        serializer = AddRemoveFavoriteSerializer()
        return Response(serializer.data)

'''Charg Wallet API'''
class ChargeWalletView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    def post(self, request):
        serializer = WalletSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            email = serializer.validated_data['email']
            try:
                customer = Customer.objects.get(email=email)
            except Customer.DoesNotExist:
                return Response("There is not any customer with the given email" , status=status.HTTP_404_NOT_FOUND)
            customer.wallet_balance += serializer.validated_data['amount']
            customer.save()
            return Response({'email' : customer.email, 'wallet_balance':customer.wallet_balance}, status=status.HTTP_200_OK)
        return Response("error!", status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        serializer = WalletSerializer()
        return Response(serializer.data)
    
'''Withdraw API'''
class WithdrawFromWalletView(APIView):
    authentication_classes = [JWTAuthentication] 
    permission_classes = [IsAuthenticated]
    def post(self, request):
        serializer = WalletSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            email = serializer.validated_data['email']
            try:
                customer = Customer.objects.get(email=email)
            except Customer.DoesNotExist:
                return Response("There is not any customer with the given email" , status=status.HTTP_404_NOT_FOUND)
            if customer.wallet_balance - serializer.validated_data['amount'] >= 0:
                customer.wallet_balance -= serializer.validated_data['amount']
                customer.save()
            else:
                return Response("The wallet balance is insufficient" , status=status.HTTP_404_NOT_FOUND)
            return Response({'email' : customer.email, 'wallet_balance':customer.wallet_balance}, status=status.HTTP_200_OK)
        return Response("error!", status=status.HTTP_400_BAD_REQUEST)
    def get(self, request):
        serializer = WalletSerializer()
        return Response(serializer.data)

'''show countries'''
class ShowAllCountry(APIView):
    def get(self, request):
        datas = Country.objects.all()
        serializer = CountrySerializer(datas, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

'''show cities of countries'''
class CitiesOfCountry(APIView):
    def post(self, request):
        country_name = request.data['name']
        country_id = Country.objects.get(name = country_name)
        cities = City.objects.filter(country_id = country_id)
        serializer = CitySerializer(cities, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    def get(self, request):
        serializer = CountrySerializer()
        return Response(serializer.data, status=status.HTTP_200_OK)
    
'''Map API'''
class LatLongUpdateRetreive(generics.RetrieveUpdateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = LatLongSerializer
    lookup_field = 'myauthor_ptr_id'
    lookup_url_kwarg = 'user_id'
    def get_queryset(self):
        print(self.kwargs)
        return Customer.objects.filter(id=self.kwargs['user_id'])

    def get_serializer_context(self):
        print(self.kwargs)
        return {'user_id': self.kwargs['user_id']}

    def patch(self, request, user_id):
        instance = get_object_or_404(Customer,myauthor_ptr_id = user_id)
        for key , value in request.data.items():
            setattr(instance,key,value)
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)   

'''get lat and long of user''' 
def get_lat_long(user_id):
    cust = get_object_or_404(Customer,id = user_id)
    content = JSONRenderer().render({'lat':cust.lat,'long':cust.lon})
    return HttpResponse(content, content_type='application/json')

'''download restaurants informations as excel in admin panel'''
class RestaurantInfoExportExcel(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="restaurants-info.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ['Name', 'Type', 'Address', 'Discount', 'Rate', 'Number', 'Manager Name', 'Manager Email']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid") 
        restaurants = Restaurant.objects.all()
        for row_num, res in enumerate(restaurants, 2):
            ws.cell(row=row_num, column=1, value=res.name)
            ws.cell(row=row_num, column=2, value=res.type)
            ws.cell(row=row_num, column=3, value=res.address)
            ws.cell(row=row_num, column=4, value=res.discount)
            ws.cell(row=row_num, column=5, value=res.rate)
            ws.cell(row=row_num, column=6, value=res.number)
            ws.cell(row=row_num, column=7, value=res.manager.name)
            ws.cell(row=row_num, column=8, value=res.manager.email)
        wb.save(response)
        return response

class OrderViewSet2(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = OrderSerializer2
    
    def get_queryset(self):
        user = self.request.user
        customer = Customer.objects.get(myauthor_ptr_id=user.id)
        return Order2.objects.filter(customer=customer)

class OrderItemViewSet2(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = OrderItemSerializer2
    
    def get_queryset(self):
        user = self.request.user
        customer = Customer.objects.get(myauthor_ptr_id=user.id)
        return OrderItem2.objects.filter(order__customer=customer)

class GetRestaurants(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        if request.user.is_admin:
            restaurants = Restaurant.objects.all()
            serializer = RestaurantSerializer(restaurants, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({'detail': 'user does not have admin permissions!'}, status=status.HTTP_401_UNAUTHORIZED)

class GetCustomers(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        if request.user.is_admin:
            customers = Customer.objects.all()
            serializer = CustomerSerializer(customers, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({'detail': 'user does not have admin permissions!'}, status=status.HTTP_401_UNAUTHORIZED)

# class TempManagerConfirmation(mixins.CreateModelMixin,mixins.DestroyModelMixin,generics.GenericAPIView):
#     serializer_class = TempManagerSerializer
#     # lookup_field = 'id'
#     def get_serializer_class(self):
#         if self.request.method == "POST":
#             return MyAuthorSerializer
#         elif self.request.method == "DELETE":
#             return TempManager.objects.all()
#         return TempManager.objects.all()
        
#     def get_queryset(self):
#         if self.request.method == "POST":
#             return MyAuthor.objects.all()
#         elif self.request.method == "DELETE":
#             return TempManager.objects.all()
#         return TempManager.objects.all()
    
#     def get_object(self):
#         queryset = self.get_queryset()
#         obj = queryset.filter(id = self.request.user.id)
#         self.check_object_permissions(self.request, obj)
#         return obj
    
#     def post(self,request,*args, **kwargs):
#         return self.create(request,*args,**kwargs)
    
#     def delete(self,request,*args, **kwargs):
#         return self.destroy(request, *args, **kwargs)


# def accept_tempMNG(request,*args, **kwargs):
#     print()

class TempManagerConfirmation(mixins.CreateModelMixin,generics.GenericAPIView):
    serializer_class = TempManagerSerializer
    # def get_serializer_class(self):
    #     if self.request.method == "POST":
    #         return MyAuthorSerializer
    #     elif self.request.method == "DELETE":
    #         return TempManager.objects.all()
    #     return TempManager.objects.all()
    
    def post(self,request,*args, **kwargs):
        serializer = TempManagerSerializer(data = request.data)
        if serializer.is_valid(raise_exception=True):
            name = request.data["name"]
            email = serializer.validated_data["email"]
            passwd = "None"
            if TempManager.objects.filter(email = email,name = name).exists():
                tmp = TempManager.objects.get(email = email,name = name)
                template = render_to_string('confirm_admin.html', {'name': name})
                data = {'to_email': email, 'body': template, 'subject': 'Your request for NoWaste has been accepted :)'}
                Util.send_email(data)
                tmp.delete()
                passwd = tmp.password
            
        # serializer = MyAuthorSerializer(data=request.data)
            # new_MyAuthor = MyAuthor.objects.create(email = email,name = name,password = passwd,role = "restaurant")
            # new_MyAuthor.save()
            new_manager = RestaurantManager.objects.create(email = email,name = name,password = passwd,role = "restaurant")
            new_manager.save()
            return Response(serializer.validated_data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


# # class TempManagerRejection(mixins.DestroyModelMixin,generics.GenericAPIView,mixins.RetrieveModelMixin):
# #     serializer_class = TempManagerSerializer
# #     queryset = TempManager.objects.all()
# #     lookup_field = 'id'
# #     lookup_url_kwarg = 'pk'
# #     def get(self,request, *args, **kwargs):
# #         return self.retrieve(request, *args, **kwargs)

# #     # def get_object(self):
# #     #     return TempManager.objects.get
# #     def destroy(request, *args, **kwargs):
# #         instance = get_object_or_404(TempManager,email = request.data['email'])
# #         instance.delete()
# #         return Response(status=status.HTTP_204_NO_CONTENT)
# class TempManagerRejection(generics.DestroyAPIView,generics.RetrieveAPIView):
#     serializer_class = TempManagerSerializer
#     queryset = TempManager.objects.all()
#     lookup_field = 'id'
#     lookup_url_kwarg = 'pk'
#     # def get(self,request, *args, **kwargs):
#     #     return self.retrieve(request, *args, **kwargs)

#     # def get_object(self):
#     #     return TempManager.objects.get
#     def destroy(request, *args, **kwargs):
#         instance = get_object_or_404(TempManager,email = request.data['email'])
#         instance.delete()
#         return Response(status=status.HTTP_204_NO_CONTENT)
class TempManagerRejection(generics.DestroyAPIView,generics.RetrieveAPIView):
    serializer_class = TempManagerSerializer
    queryset = TempManager.objects.all()
    lookup_field = 'id'
    lookup_url_kwarg = 'pk'
    def get(self,request, *args, **kwargs):
        pk = self.kwargs['pk']
        if TempManager.objects.filter(id = pk).exists():
            tmp = TempManager.objects.get(id = pk)
            email = tmp.email
            name = tmp.name
            template = render_to_string('reject_admin.html', {'name': name})
            data = {'to_email': email, 'body': template, 'subject': 'Your request for NoWaste has been rejected :('}
            Util.send_email(data)
            tmp.delete()
        return Response("User rejected and email sent.", status=status.HTTP_200_OK)
    # def get(self,request, *args, **kwargs):
    #     return self.retrieve(request, *args, **kwargs)

    # # def get_object(self):
    # #     return TempManager.objects.get
    # def destroy(request, *args, **kwargs):
    #     instance = get_object_or_404(TempManager,email = request.data['email'])
    #     instance.delete()
    #     return Response(status=status.HTTP_204_NO_CONTENT)



class AdminProfile(generics.ListAPIView):
    serializer_class = ManagerSerialzer
    queryset = RestaurantManager.objects.all()
